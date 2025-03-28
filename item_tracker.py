import tkinter as tk
from tkinter import messagebox, colorchooser, font as tkFont, ttk, filedialog
from openpyxl import Workbook, load_workbook
import configparser
import os
import keyboard  # For global keybinds

class HuntShowdownGunathonTracker:
    def __init__(self, root):
        self.root = root
        self.root.title("Hunt Showdown Gunathon Tracker")
        self.items = []
        self.current_kills = 0
        self.overlay_visible = True
        self.overlay_color = "purple"
        self.highlight_color = None
        self.font_style = {"bold": False, "italic": False, "underline": False}
        self.active_weapon_index = 0
        self.font_family = "Unfair Style 2 Rough"
        self.font_size = 16
        self.compact_mode = False
        self.overlay_custom_text = {
            "kills": "Kills: {kills}",
            "weapon": "Active Weapon: {weapon}",
            "custom": "",
        }

        self.keybinds = {
            "increase_1": "<Up>",
            "decrease_1": "<Down>",
            "toggle_overlay": "ctrl+o",
        }

        self.config = configparser.ConfigParser()
        self.load_preferences()
        self.load_items_from_excel("items.xlsx")

        self.create_menu()

        # Main UI Elements
        self.kills_label = tk.Label(root, text="Kills: 0", font=("Arial", 14))
        self.kills_label.pack(pady=10)

        self.active_weapon_label = tk.Label(root, text="Active Weapon: None", font=("Arial", 14))
        self.active_weapon_label.pack(pady=5)

        self.kill_adjust_frame = tk.Frame(root)
        self.kill_adjust_frame.pack(pady=5)

        self.increase_by_1_button = tk.Button(self.kill_adjust_frame, text="+1", command=lambda: self.adjust_kills(1))
        self.increase_by_1_button.pack(side=tk.LEFT, padx=5)

        self.increase_by_2_button = tk.Button(self.kill_adjust_frame, text="+2", command=lambda: self.adjust_kills(2))
        self.increase_by_2_button.pack(side=tk.LEFT, padx=5)

        self.decrease_by_1_button = tk.Button(self.kill_adjust_frame, text="-1", command=lambda: self.adjust_kills(-1))
        self.decrease_by_1_button.pack(side=tk.LEFT, padx=5)

        self.decrease_by_2_button = tk.Button(self.kill_adjust_frame, text="-2", command=lambda: self.adjust_kills(-2))
        self.decrease_by_2_button.pack(side=tk.LEFT, padx=5)

        self.custom_adjust_frame = tk.Frame(root)
        self.custom_adjust_frame.pack(pady=5)

        self.custom_adjust_label = tk.Label(self.custom_adjust_frame, text="Adjust by:", font=("Arial", 12))
        self.custom_adjust_label.pack(side=tk.LEFT, padx=5)

        self.custom_adjust_entry = tk.Entry(self.custom_adjust_frame, width=10, font=("Arial", 12))
        self.custom_adjust_entry.pack(side=tk.LEFT, padx=5)

        self.custom_adjust_button = tk.Button(self.custom_adjust_frame, text="Apply", command=self.adjust_kills_custom)
        self.custom_adjust_button.pack(side=tk.LEFT, padx=5)

        self.restart_button = tk.Button(root, text="Restart Gunathon", command=self.restart_gunathon)
        self.restart_button.pack(pady=10)

        self.toggle_overlay_button = tk.Button(root, text="Hide Overlay", command=self.toggle_overlay)
        self.toggle_overlay_button.pack(pady=10)

        self.overlay_settings_button = tk.Button(root, text="Overlay Settings", command=self.open_overlay_settings)
        self.overlay_settings_button.pack(pady=5)

        self.create_overlay()
        self.setup_keybinds()
        self.update_ui()

    def create_menu(self):
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        weapons_menu = tk.Menu(menubar, tearoff=0)
        weapons_menu.add_command(label="View All Weapons", command=self.view_all_weapons)
        weapons_menu.add_command(label="Add New Weapon", command=self.add_new_weapon)
        weapons_menu.add_separator()
        weapons_menu.add_command(label="Export Weapons to Excel", command=self.export_items_to_excel)
        weapons_menu.add_command(label="Import Weapons from Excel", command=self.import_items_from_excel)
        menubar.add_cascade(label="Weapons", menu=weapons_menu)

        keybinds_menu = tk.Menu(menubar, tearoff=0)
        keybinds_menu.add_command(label="Customize Keybinds", command=self.customize_keybinds)
        menubar.add_cascade(label="Keybinds", menu=keybinds_menu)

    def load_items_from_excel(self, filename="items.xlsx"):
        if not os.path.exists(filename):
            self.create_default_excel(filename)
        
        try:
            workbook = load_workbook(filename)
            sheet = workbook.active
            
            if sheet.max_row > 1:
                kill_cell = sheet.cell(row=2, column=5)
                self.current_kills = int(kill_cell.value) if kill_cell.value is not None else 0
                
                index_cell = sheet.cell(row=2, column=6)
                self.active_weapon_index = int(index_cell.value) if index_cell.value is not None else 0
            
            self.items = []
            for row in sheet.iter_rows(min_row=3, values_only=True):
                if row[0]:
                    self.items.append({
                        "weapon": row[0],
                        "range_start": int(row[1]),
                        "range_end": int(row[2]),
                        "status": row[3] if len(row) > 3 else "Incomplete"
                    })
                    
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load items: {str(e)}\nCreating new default file.")
            self.create_default_excel(filename)
            self.load_items_from_excel(filename)

    def create_default_excel(self, filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Weapons"
        ws.append(["Weapon", "Range Start", "Range End", "Status", "Kills", "Active Weapon Index"])
        ws.append([None, None, None, None, 0, 0])
        
        weapons = [
            ["Nagant", -999, 5, "Complete"],
            ["Nagant Precision", 5, 10, "Active"],
            ["Nagant Silencer", 10, 15, "Complete"],
            ["Nagant Precision Deadeye", 15, 20, "Complete"],
            ["Nagant Officer", 20, 25, "Complete"],
            ["Nagant Officer Brawler", 25, 30, "Complete"],
            ["Nagant Carbine", 30, 35, "Complete"],
            ["Nagant Carbine Deadeye", 35, 40, "Complete"],
            ["Romero", 40, 45, "Complete"],
            ["Romero Handcannon", 45, 50, "Active"],
            ["Romero Talon", 50, 55, "Incomplete"],
            ["Romero Hatchet", 55, 60, "Incomplete"],
            ["Romero Alamo", 60, 65, "Incomplete"],
            ["Scottfield", 65, 70, "Incomplete"],
            ["Scottfield Brawler", 70, 75, "Incomplete"],
            ["Scottfield Spitfire", 75, 80, "Incomplete"],
            ["Scottfield Precision", 80, 85, "Incomplete"],
            ["Scottfield Swift", 85, 90, "Incomplete"],
            ["Springfield", 90, 95, "Incomplete"],
            ["Springfield Marksman", 95, 100, "Incomplete"],
            ["Springfield Compact", 100, 105, "Incomplete"],
            ["Springfield Compact Striker", 105, 110, "Incomplete"],
            ["Springfield Compact Deadeye", 110, 115, "Incomplete"],
            ["Springfield Bayonet", 115, 120, "Incomplete"],
            ["1865 Carbine", 120, 125, "Incomplete"],
            ["1865 Carbine Aperture", 125, 130, "Incomplete"],
            ["Frontier 73C", 130, 135, "Incomplete"],
            ["Ranger 73C Silencer", 135, 140, "Incomplete"],
            ["Ranger 73C Marksman", 140, 145, "Incomplete"],
            ["Ranger 73C Vandal", 145, 150, "Incomplete"],
            ["Ranger 73C Striker", 150, 155, "Incomplete"],
            ["Ranger 73C Deadeye", 155, 160, "Incomplete"],
            ["Terminus", 160, 165, "Incomplete"],
            ["Terminus Handcannon", 165, 170, "Incomplete"],
            ["Martini", 170, 175, "Incomplete"],
            ["Martini Deadeye", 175, 180, "Incomplete"],
            ["Martini Riposte", 180, 185, "Incomplete"],
            ["Martini Marksman", 185, 190, "Incomplete"],
            ["Martini Ironside", 190, 195, "Incomplete"],
            ["Caldwell New Army", 195, 200, "Incomplete"],
            ["Caldwell New Army Swift", 200, 205, "Incomplete"],
            ["Vetterli", 205, 210, "Incomplete"],
            ["Vetterli Deadeye", 210, 215, "Incomplete"],
            ["Vetterli Marksman", 215, 220, "Incomplete"],
            ["Vetterli Bayonet", 220, 225, "Incomplete"],
            ["Vetterli Silencer", 225, 230, "Incomplete"],
            ["Vetterli Cyclone", 230, 235, "Incomplete"],
            ["Caldwell Rival", 235, 240, "Incomplete"],
            ["Caldwell Rival Trauma", 240, 245, "Incomplete"],
            ["Caldwell Rival Handcannon", 245, 250, "Incomplete"],
            ["Caldwell Rival Mace", 250, 255, "Incomplete"],
            ["Conversion Pistol", 255, 260, "Incomplete"],
            ["Conversion Chain", 260, 265, "Incomplete"],
            ["Conversion Uppercut", 265, 270, "Incomplete"],
            ["Uppercut Precision", 270, 275, "Incomplete"],
            ["Uppercut Precision Deadeye", 275, 280, "Incomplete"],
            ["Hand Crossbow", 280, 285, "Incomplete"],
            ["Chu Ko Nu", 285, 290, "Incomplete"],
            ["Crossbow", 290, 295, "Incomplete"],
            ["Crossbow Deadeye", 295, 300, "Incomplete"],
            ["Centennial", 300, 305, "Incomplete"],
            ["Centennial Shorty", 305, 310, "Incomplete"],
            ["Centennial Sniper", 310, 315, "Incomplete"],
            ["Centennial Shorty Silencer", 315, 320, "Incomplete"],
            ["Centennial Trauma", 320, 325, "Incomplete"],
            ["Centennial Pointman", 325, 330, "Incomplete"],
            ["Mako", 330, 335, "Incomplete"],
            ["Mako Aperture", 335, 340, "Incomplete"],
            ["Mako Claw", 340, 345, "Incomplete"],
            ["Ranger", 345, 350, "Incomplete"],
            ["Winfield Aperture", 350, 355, "Incomplete"],
            ["Winfield Talon", 355, 360, "Incomplete"],
            ["Winfield Swift", 360, 365, "Incomplete"],
            ["Winfield Bayonet", 365, 370, "Incomplete"],
            ["Maynard", 370, 375, "Incomplete"],
            ["Maynard Sniper", 375, 380, "Incomplete"],
            ["Maynard Sniper Silencer", 380, 385, "Incomplete"],
            ["Caldwell Pax", 385, 390, "Incomplete"],
            ["Pax Trueshot", 390, 395, "Incomplete"],
            ["Pax Claw", 395, 400, "Incomplete"],
            ["Marathon", 400, 405, "Incomplete"],
            ["Marathon Swift", 405, 410, "Incomplete"],
            ["Sparks", 410, 415, "Incomplete"],
            ["Sparks Pistol", 415, 420, "Incomplete"],
            ["Sparks Pistol Silencer", 420, 425, "Incomplete"],
            ["Sparks Silencer", 425, 430, "Incomplete"],
            ["Sparks Sniper", 430, 435, "Incomplete"],
            ["Bornheim", 435, 440, "Incomplete"],
            ["Bornheim Match", 440, 445, "Incomplete"],
            ["Bornheim Silencer", 445, 450, "Incomplete"],
            ["Bornheim Extended", 450, 455, "Incomplete"],
            ["Specter", 455, 460, "Incomplete"],
            ["Specter Compact", 460, 465, "Incomplete"],
            ["Specter Bayonet", 465, 470, "Incomplete"],
            ["LeMat", 470, 475, "Incomplete"],
            ["LeMat Carbine", 475, 480, "Incomplete"],
            ["LeMat Carbine Marksman", 480, 485, "Incomplete"],
            ["Uppermat (Haymaker)", 485, 490, "Incomplete"],
            ["Berthier", 490, 495, "Incomplete"],
            ["Berthier Reposte", 495, 500, "Incomplete"],
            ["Berthier Deadeye", 500, 505, "Incomplete"],
            ["Berthier Markman", 505, 510, "Incomplete"],
            ["Infantry 73L", 510, 515, "Incomplete"],
            ["Infantry Bayonet", 515, 520, "Incomplete"],
            ["Infantry Sniper", 520, 525, "Incomplete"],
            ["Lebel", 525, 530, "Incomplete"],
            ["Lebel Aperture", 530, 535, "Incomplete"],
            ["Lebel Talon", 535, 540, "Incomplete"],
            ["Lebel Marksman", 540, 545, "Incomplete"],
            ["Hunting Bow", 545, 550, "Incomplete"],
            ["Drilling", 550, 555, "Incomplete"],
            ["Drilling Shorty", 555, 560, "Incomplete"],
            ["Drilling Hatchet", 560, 565, "Incomplete"],
            ["Krag", 565, 570, "Incomplete"],
            ["Krag Bayonet", 570, 575, "Incomplete"],
            ["Krag Silencer", 575, 580, "Incomplete"],
            ["Krag Sniper", 580, 585, "Incomplete"],
            ["Slate", 585, 590, "Incomplete"],
            ["Slate Reposte", 590, 595, "Incomplete"],
            ["Dolch", 595, 600, "Incomplete"],
            ["Dolch Claw", 600, 605, "Incomplete"],
            ["Dolch Precision", 605, 610, "Incomplete"],
            ["Dolch Deadeye", 610, 615, "Incomplete"],
            ["Mosin", 615, 620, "Incomplete"],
            ["Mosin Obrez", 620, 625, "Incomplete"],
            ["Mosin Bayonet", 625, 630, "Incomplete"],
            ["Mosin Mace", 630, 635, "Incomplete"],
            ["Mosin Sniper", 635, 640, "Incomplete"],
            ["Mosin Drum", 640, 645, "Incomplete"],
            ["Avtomat", 645, 650, "Incomplete"],
            ["Bomb Launcher", 650, 655, "Incomplete"],
            ["Bomblance", 655, 660, "Incomplete"],
            ["Auto-5", 660, 665, "Incomplete"],
            ["Auto-4 Shorty", 665, 670, "Incomplete"],
            ["Nitro", 670, 675, "Incomplete"],
            ["Derringer", 675, 680, "Incomplete"],
            ["Knife", 680, 685, "Incomplete"],
            ["Baseball Bat", 685, 690, "Incomplete"],
            ["Cavalry Saber", 690, 695, "Incomplete"],
            ["Combat Axe", 695, 700, "Incomplete"],
            ["Katana", 700, 705, "Incomplete"],
            ["Machete", 705, 710, "Incomplete"],
            ["Railroad Hammer", 710, 715, "Incomplete"]
        ]
        
        for weapon in weapons:
            ws.append(weapon)
        
        wb.save(filename)
        messagebox.showinfo("New File Created", f"Created new weapons file: {filename}")

    def view_all_weapons(self):
        weapons_window = tk.Toplevel(self.root)
        weapons_window.title("All Weapons")

        for item in self.items:
            status = item["status"]
            weapon_label = tk.Label(
                weapons_window,
                text=f"{item['weapon']} (Range: {item['range_start']}-{item['range_end']}) - {status}",
                font=("Arial", 12),
            )
            weapon_label.pack(pady=5)

    def add_new_weapon(self):
        add_weapon_window = tk.Toplevel(self.root)
        add_weapon_window.title("Add New Weapon")

        tk.Label(add_weapon_window, text="Weapon Name:", font=("Arial", 12)).pack(pady=5)
        weapon_name_entry = tk.Entry(add_weapon_window, font=("Arial", 12))
        weapon_name_entry.pack(pady=5)

        tk.Label(add_weapon_window, text="Range Start:", font=("Arial", 12)).pack(pady=5)
        range_start_entry = tk.Entry(add_weapon_window, font=("Arial", 12))
        range_start_entry.pack(pady=5)

        tk.Label(add_weapon_window, text="Range End:", font=("Arial", 12)).pack(pady=5)
        range_end_entry = tk.Entry(add_weapon_window, font=("Arial", 12))
        range_end_entry.pack(pady=5)

        add_button = tk.Button(
            add_weapon_window,
            text="Add Weapon",
            command=lambda: self.save_new_weapon(
                weapon_name_entry.get(),
                range_start_entry.get(),
                range_end_entry.get(),
                add_weapon_window,
            ),
        )
        add_button.pack(pady=10)

    def save_new_weapon(self, name, range_start, range_end, window):
        try:
            range_start = int(range_start)
            range_end = int(range_end)
            if range_start >= range_end:
                raise ValueError("Range start must be less than range end.")
        except ValueError as e:
            messagebox.showerror("Invalid Input", str(e))
            return

        self.items.append({
            "weapon": name,
            "range_start": range_start,
            "range_end": range_end,
            "status": "Incomplete",
        })

        self.update_ui()
        window.destroy()
        messagebox.showinfo("Success", "Weapon added successfully!")

    def export_items_to_excel(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save Weapons List As"
        )
        if file_path:
            try:
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(["Weapon", "Range Start", "Range End", "Status", "Kills", "Active Weapon Index"])
                sheet.append([None, None, None, None, self.current_kills, self.active_weapon_index])
                
                for item in self.items:
                    sheet.append([item["weapon"], item["range_start"], item["range_end"], item["status"]])
                
                workbook.save(file_path)
                messagebox.showinfo("Success", "Weapons list exported successfully!")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to export weapons list: {str(e)}")

    def import_items_from_excel(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Open Weapons List"
        )
        if file_path:
            try:
                workbook = load_workbook(file_path)
                sheet = workbook.active
                self.items.clear()

                if sheet.max_row > 1:
                    self.current_kills = int(sheet.cell(row=2, column=5).value or 0)
                    self.active_weapon_index = int(sheet.cell(row=2, column=6).value or 0)

                for row in sheet.iter_rows(min_row=3, values_only=True):
                    if row[0]:
                        self.items.append({
                            "weapon": row[0],
                            "range_start": int(row[1]),
                            "range_end": int(row[2]),
                            "status": row[3] if len(row) > 3 else "Incomplete"
                        })

                self.update_ui()
                messagebox.showinfo("Success", "Weapons list imported successfully!")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to import weapons list: {str(e)}")

    def setup_keybinds(self):
        for key in self.keybinds.values():
            self.root.unbind(key)

        self.root.bind(self.keybinds["increase_1"], lambda event: self.adjust_kills(1))
        self.root.bind(self.keybinds["decrease_1"], lambda event: self.adjust_kills(-1))
        self.root.bind(self.keybinds["toggle_overlay"], lambda event: self.toggle_overlay())

        try:
            keyboard.unhook_all()
            keyboard.add_hotkey(self.keybinds["increase_1"].strip("<>"), lambda: self.adjust_kills(1))
            keyboard.add_hotkey(self.keybinds["decrease_1"].strip("<>"), lambda: self.adjust_kills(-1))
            keyboard.add_hotkey("ctrl+o", self.toggle_overlay)
        except:
            pass

    def customize_keybinds(self):
        keybind_window = tk.Toplevel(self.root)
        keybind_window.title("Customize Keybinds")

        tk.Label(keybind_window, text="Increase Kills by 1:", font=("Arial", 12)).pack(pady=5)
        increase_1_entry = tk.Entry(keybind_window, font=("Arial", 12))
        increase_1_entry.insert(0, self.keybinds["increase_1"])
        increase_1_entry.pack(pady=5)

        tk.Label(keybind_window, text="Decrease Kills by 1:", font=("Arial", 12)).pack(pady=5)
        decrease_1_entry = tk.Entry(keybind_window, font=("Arial", 12))
        decrease_1_entry.insert(0, self.keybinds["decrease_1"])
        decrease_1_entry.pack(pady=5)

        tk.Label(keybind_window, text="Toggle Overlay:", font=("Arial", 12)).pack(pady=5)
        toggle_overlay_entry = tk.Entry(keybind_window, font=("Arial", 12))
        toggle_overlay_entry.insert(0, self.keybinds["toggle_overlay"])
        toggle_overlay_entry.pack(pady=5)

        save_button = tk.Button(
            keybind_window,
            text="Save Keybinds",
            command=lambda: self.save_keybinds(
                increase_1_entry.get(),
                decrease_1_entry.get(),
                toggle_overlay_entry.get(),
                keybind_window,
            ),
        )
        save_button.pack(pady=10)

    def save_keybinds(self, increase_1, decrease_1, toggle_overlay, window):
        try:
            self.root.bind(increase_1, lambda event: None)
            self.root.bind(decrease_1, lambda event: None)
            self.root.bind(toggle_overlay, lambda event: None)
        except tk.TclError:
            messagebox.showerror("Invalid Keybind", "One or more keybinds are invalid.")
            return

        self.keybinds = {
            "increase_1": increase_1,
            "decrease_1": decrease_1,
            "toggle_overlay": toggle_overlay,
        }

        self.setup_keybinds()
        window.destroy()
        messagebox.showinfo("Success", "Keybinds updated successfully!")

    def toggle_compact_mode(self):
        self.compact_mode = not self.compact_mode
        if self.compact_mode:
            self.overlay.geometry("200x50")
            self.overlay_kills_label.pack_forget()
            self.overlay_active_label.pack_forget()
        else:
            self.overlay.geometry("400x150")
            self.overlay_kills_label.pack(pady=5)
            self.overlay_active_label.pack(pady=5)

    def load_preferences(self):
        if os.path.exists("preferences.ini"):
            self.config.read("preferences.ini")
            if "Preferences" in self.config:
                self.overlay_color = self.config["Preferences"].get("overlay_color", "purple")
                self.highlight_color = self.config["Preferences"].get("highlight_color", None)
                self.font_family = self.config["Preferences"].get("font_family", "Unfair Style 2 Rough")
                self.font_size = int(self.config["Preferences"].get("font_size", 16))
                self.font_style["bold"] = self.config["Preferences"].getboolean("bold", False)
                self.font_style["italic"] = self.config["Preferences"].getboolean("italic", False)
                self.font_style["underline"] = self.config["Preferences"].getboolean("underline", False)
                self.keybinds = {
                    "increase_1": self.config["Preferences"].get("increase_1", "<Up>"),
                    "decrease_1": self.config["Preferences"].get("decrease_1", "<Down>"),
                    "toggle_overlay": self.config["Preferences"].get("toggle_overlay", "ctrl+o"),
                }
                self.compact_mode = self.config["Preferences"].getboolean("compact_mode", False)
                self.overlay_custom_text["kills"] = self.config["Preferences"].get("kills_text", "Kills: {kills}")
                self.overlay_custom_text["weapon"] = self.config["Preferences"].get("weapon_text", "Active Weapon: {weapon}")
                self.overlay_custom_text["custom"] = self.config["Preferences"].get("custom_text", "")

    def save_preferences(self):
        self.config["Preferences"] = {
            "overlay_color": self.overlay_color,
            "highlight_color": self.highlight_color if self.highlight_color else "",
            "font_family": self.font_family,
            "font_size": str(self.font_size),
            "bold": str(self.font_style["bold"]),
            "italic": str(self.font_style["italic"]),
            "underline": str(self.font_style["underline"]),
            "increase_1": self.keybinds["increase_1"],
            "decrease_1": self.keybinds["decrease_1"],
            "toggle_overlay": self.keybinds["toggle_overlay"],
            "compact_mode": str(self.compact_mode),
            "kills_text": self.overlay_custom_text["kills"],
            "weapon_text": self.overlay_custom_text["weapon"],
            "custom_text": self.overlay_custom_text["custom"],
        }
        with open("preferences.ini", "w") as configfile:
            self.config.write(configfile)

    def create_overlay(self):
        self.overlay = tk.Toplevel(self.root)
        self.overlay.title("Overlay - Hunt Showdown Gunathon Tracker")
        self.overlay.geometry("400x150+100+100")
        self.overlay.attributes("-alpha", 0.8)
        self.overlay.configure(bg="black")
        self.overlay.attributes("-transparentcolor", "black")
        self.overlay.protocol("WM_DELETE_WINDOW", self.on_overlay_close)

        self.overlay_kills_label = self.create_overlay_label("", pady=5)
        self.overlay_weapon_label = self.create_overlay_label("", pady=5)
        self.overlay_custom_label = self.create_overlay_label("", pady=5)

        if self.compact_mode:
            self.toggle_compact_mode()

        self.overlay.update_idletasks()
        self.overlay.geometry(f"{self.overlay.winfo_width()}x{self.overlay.winfo_height()}")

    def create_overlay_label(self, text, pady=0):
        label = tk.Label(
            self.overlay,
            text=text,
            font=self.get_font(),
            fg=self.overlay_color,
            bg=self.highlight_color if self.highlight_color else "black",
        )
        label.pack(pady=pady)
        return label

    def on_overlay_close(self):
        self.toggle_overlay()

    def get_font(self):
        return tkFont.Font(
            family=self.font_family,
            size=self.font_size,
            weight="bold" if self.font_style["bold"] else "normal",
            slant="italic" if self.font_style["italic"] else "roman",
            underline=self.font_style["underline"],
        )

    def toggle_font_style(self, style):
        self.font_style[style] = not self.font_style[style]
        self.update_ui()

    def change_overlay_color(self):
        color = colorchooser.askcolor(title="Choose Overlay Text Color")
        if color[1]:
            self.overlay_color = color[1]
            self.update_ui()

    def change_highlight_color(self):
        color = colorchooser.askcolor(title="Choose Highlight Color")
        if color[1]:
            self.highlight_color = color[1]
            self.update_ui()

    def change_font(self):
        font_window = tk.Toplevel(self.root)
        font_window.title("Select Font")

        font_families = list(tkFont.families())
        font_families.sort()

        tk.Label(font_window, text="Select Font Family:", font=("Arial", 12)).pack(pady=5)
        font_combobox = ttk.Combobox(font_window, values=font_families, font=("Arial", 12))
        font_combobox.pack(pady=5)
        font_combobox.set(self.font_family)

        apply_button = tk.Button(
            font_window,
            text="Apply",
            command=lambda: self.apply_font(font_combobox.get(), font_window),
        )
        apply_button.pack(pady=10)

    def apply_font(self, font_family, window):
        self.font_family = font_family
        self.update_ui()
        window.destroy()

    def adjust_kills_custom(self):
        try:
            amount = int(self.custom_adjust_entry.get())
            self.adjust_kills(amount)
        except ValueError:
            messagebox.showerror("Invalid Input", "Please enter a valid number.")

    def restart_gunathon(self):
        self.current_kills = 0
        for index, item in enumerate(self.items):
            if index == 0:
                item["status"] = "Active"
            else:
                item["status"] = "Incomplete"
        self.update_ui()

    def update_ui(self):
        self.kills_label.config(text=f"Kills: {self.current_kills}")
        active_weapon = self.get_active_weapon()
        self.active_weapon_label.config(text=f"Active Weapon: {active_weapon['weapon']}" if active_weapon else "Active Weapon: None")

        self.overlay_kills_label.config(
            text=self.add_letter_spacing(self.overlay_custom_text["kills"].format(kills=self.current_kills)),
            font=self.get_font(),
            fg=self.overlay_color,
            bg=self.highlight_color if self.highlight_color else "black",
        )
        self.overlay_weapon_label.config(
            text=self.add_letter_spacing(self.overlay_custom_text["weapon"].format(weapon=active_weapon["weapon"] if active_weapon else "None")),
            font=self.get_font(),
            fg=self.overlay_color,
            bg=self.highlight_color if self.highlight_color else "black",
        )
        self.overlay_custom_label.config(
            text=self.add_letter_spacing(self.overlay_custom_text["custom"]),
            font=self.get_font(),
            fg=self.overlay_color,
            bg=self.highlight_color if self.highlight_color else "black",
        )

        self.resize_overlay()

    def resize_overlay(self):
        width = int(self.font_size * 20)
        height = int(self.font_size * 10)
        self.overlay.geometry(f"{width}x{height}")

    def add_letter_spacing(self, text, spacing=1):
        return " ".join(text)

    def toggle_overlay(self):
        if self.overlay_visible:
            self.overlay.withdraw()
            self.toggle_overlay_button.config(text="Show Overlay")
        else:
            self.overlay.deiconify()
            self.toggle_overlay_button.config(text="Hide Overlay")
        self.overlay_visible = not self.overlay_visible

    def adjust_kills(self, amount):
        self.current_kills += amount
        self.update_active_weapon()
        self.update_ui()

    def update_active_weapon(self):
        for item in self.items:
            if self.current_kills < item["range_start"]:
                item["status"] = "Incomplete"
            elif self.current_kills >= item["range_end"]:
                item["status"] = "Complete"
            else:
                item["status"] = "Active"
                self.active_weapon_index = self.items.index(item)

    def get_active_weapon(self):
        for item in self.items:
            if item["status"] == "Active":
                return item
        return None

    def on_closing(self):
        self.save_preferences()
        self.save_items_to_excel("items.xlsx")
        self.root.destroy()

    def open_overlay_settings(self):
        overlay_settings_window = tk.Toplevel(self.root)
        overlay_settings_window.title("Overlay Settings")

        tk.Label(overlay_settings_window, text="Kills Text Template:", font=("Arial", 12)).pack(pady=5)
        self.kills_text_entry = tk.Entry(overlay_settings_window, width=30, font=("Arial", 12))
        self.kills_text_entry.insert(0, self.overlay_custom_text["kills"])
        self.kills_text_entry.pack(pady=5)

        tk.Label(overlay_settings_window, text="Weapon Text Template:", font=("Arial", 12)).pack(pady=5)
        self.weapon_text_entry = tk.Entry(overlay_settings_window, width=30, font=("Arial", 12))
        self.weapon_text_entry.insert(0, self.overlay_custom_text["weapon"])
        self.weapon_text_entry.pack(pady=5)

        tk.Label(overlay_settings_window, text="Custom Text:", font=("Arial", 12)).pack(pady=5)
        self.custom_text_entry = tk.Entry(overlay_settings_window, width=30, font=("Arial", 12))
        self.custom_text_entry.insert(0, self.overlay_custom_text["custom"])
        self.custom_text_entry.pack(pady=5)

        apply_button = tk.Button(
            overlay_settings_window,
            text="Apply",
            command=self.update_custom_overlay_text,
        )
        apply_button.pack(pady=10)

        tk.Label(overlay_settings_window, text="Overlay Text Color:", font=("Arial", 12)).pack(pady=5)
        self.color_picker_button = tk.Button(
            overlay_settings_window,
            text="Choose Color",
            command=self.change_overlay_color,
        )
        self.color_picker_button.pack(pady=5)

        tk.Label(overlay_settings_window, text="Highlight Color:", font=("Arial", 12)).pack(pady=5)
        self.highlight_color_button = tk.Button(
            overlay_settings_window,
            text="Choose Color",
            command=self.change_highlight_color,
        )
        self.highlight_color_button.pack(pady=5)

        tk.Label(overlay_settings_window, text="Font Style:", font=("Arial", 12)).pack(pady=5)
        self.bold_toggle_button = tk.Button(
            overlay_settings_window,
            text="Toggle Bold",
            command=lambda: self.toggle_font_style("bold"),
        )
        self.bold_toggle_button.pack(pady=5)

        self.italic_toggle_button = tk.Button(
            overlay_settings_window,
            text="Toggle Italic",
            command=lambda: self.toggle_font_style("italic"),
        )
        self.italic_toggle_button.pack(pady=5)

        self.underline_toggle_button = tk.Button(
            overlay_settings_window,
            text="Toggle Underline",
            command=lambda: self.toggle_font_style("underline"),
        )
        self.underline_toggle_button.pack(pady=5)

        tk.Label(overlay_settings_window, text="Font Family:", font=("Arial", 12)).pack(pady=5)
        self.font_selector_button = tk.Button(
            overlay_settings_window,
            text="Change Font",
            command=self.change_font,
        )
        self.font_selector_button.pack(pady=5)

        tk.Label(overlay_settings_window, text="Font Size:", font=("Arial", 12)).pack(pady=5)
        self.font_size_entry = tk.Entry(overlay_settings_window, width=10, font=("Arial", 12))
        self.font_size_entry.insert(0, str(self.font_size))
        self.font_size_entry.pack(pady=5)

        apply_font_size_button = tk.Button(
            overlay_settings_window,
            text="Apply Font Size",
            command=self.apply_font_size,
        )
        apply_font_size_button.pack(pady=10)

    def apply_font_size(self):
        try:
            self.font_size = int(self.font_size_entry.get())
            self.update_ui()
        except ValueError:
            messagebox.showerror("Invalid Input", "Please enter a valid font size.")

    def update_custom_overlay_text(self):
        self.overlay_custom_text["kills"] = self.kills_text_entry.get()
        self.overlay_custom_text["weapon"] = self.weapon_text_entry.get()
        self.overlay_custom_text["custom"] = self.custom_text_entry.get()
        self.update_ui()

    def save_items_to_excel(self, filename):
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Weapon", "Range Start", "Range End", "Status", "Kills", "Active Weapon Index"])
            sheet.append([None, None, None, None, self.current_kills, self.active_weapon_index])
            
            for item in self.items:
                sheet.append([item["weapon"], item["range_start"], item["range_end"], item["status"]])
            
            workbook.save(filename)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save items: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = HuntShowdownGunathonTracker(root)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    root.mainloop()